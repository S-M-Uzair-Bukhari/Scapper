import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from lead_scraper.config import from_root


OUTPUT_FILE = from_root("data", "leads.xlsx")

LEAD_COLUMNS = [
    ("Source", "source", 14),
    ("Source Lead ID", "sourceLeadId", 22),
    ("Title", "title", 45),
    ("Description", "description", 80),
    ("Category", "category", 20),
    ("Country", "country", 20),
    ("Posted At", "postedAt", 24),
    ("Posted Raw", "postedAtRaw", 18),
    ("Scraped At", "scrapedAt", 24),
    ("Budget", "budget", 18),
    ("Score", "score", 10),
    ("Score Reasons", "scoreReasons", 45),
    ("URL", "url", 60),
    ("Dedupe Key", "dedupeKey", 66),
]

LOG_COLUMNS = [
    ("Run Time", "runTime", 24),
    ("Source", "source", 14),
    ("Category", "category", 22),
    ("Found", "found", 10),
    ("Saved", "saved", 10),
    ("Duplicates", "duplicates", 12),
    ("Skipped Old", "skippedOld", 12),
    ("Skipped Low Score", "skippedLowScore", 18),
    ("Errors", "errors", 40),
]


def with_file_retry(action, label):
    last_error = None
    for attempt in range(6):
        try:
            return action()
        except PermissionError as error:
            last_error = error
            if attempt == 5:
                break
            time.sleep(1)

    raise RuntimeError(f"{label} failed because {OUTPUT_FILE} is locked. Close the Excel file and run again.") from last_error


def ensure_sheet(workbook, name, columns):
    sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)

    for index, (header, _key, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index)
        if not cell.value:
            cell.value = header
        cell.font = Font(bold=True)
        sheet.column_dimensions[cell.column_letter].width = width

    sheet.freeze_panes = "A2"
    return sheet


def create_workbook():
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    ensure_sheet(workbook, "Priority Leads", LEAD_COLUMNS)
    ensure_sheet(workbook, "All Scraped Leads", LEAD_COLUMNS)
    ensure_sheet(workbook, "Rejected Low Score", LEAD_COLUMNS)
    ensure_sheet(workbook, "Run Logs", LOG_COLUMNS)

    return workbook


def load_or_create_workbook():
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not OUTPUT_FILE.exists():
        return create_workbook()
    return with_file_retry(lambda: load_workbook(OUTPUT_FILE), "Excel read")


def serialize_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return value if value is not None else ""


def append_dict(sheet, columns, row):
    sheet.append([serialize_value(row.get(key)) for _header, key, _width in columns])


def read_existing_lead_keys():
    workbook = load_or_create_workbook()
    rows = []

    for sheet_name in ("Priority Leads", "All Scraped Leads", "Rejected Low Score"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({
                "sourceLeadId": values[1] if len(values) > 1 else "",
                "url": values[12] if len(values) > 12 else "",
                "dedupeKey": values[13] if len(values) > 13 else "",
            })

    return rows


def save_run_result(accepted_leads, rejected_leads, logs):
    workbook = load_or_create_workbook()
    priority = ensure_sheet(workbook, "Priority Leads", LEAD_COLUMNS)
    all_sheet = ensure_sheet(workbook, "All Scraped Leads", LEAD_COLUMNS)
    rejected = ensure_sheet(workbook, "Rejected Low Score", LEAD_COLUMNS)
    run_logs = ensure_sheet(workbook, "Run Logs", LOG_COLUMNS)

    for lead in accepted_leads:
        append_dict(priority, LEAD_COLUMNS, lead)
        append_dict(all_sheet, LEAD_COLUMNS, lead)

    for lead in rejected_leads:
        append_dict(rejected, LEAD_COLUMNS, lead)
        append_dict(all_sheet, LEAD_COLUMNS, lead)

    for log in logs:
        append_dict(run_logs, LOG_COLUMNS, log)

    with_file_retry(lambda: workbook.save(OUTPUT_FILE), "Excel write")


def get_output_file():
    return OUTPUT_FILE

